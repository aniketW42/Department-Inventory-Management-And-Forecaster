import csv
import io
import datetime
import calendar
import openpyxl
from openpyxl.utils import get_column_letter

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.http import HttpResponse, HttpResponseBadRequest

from .models import ItemRequest, InventoryItem


# ------------------------ 📊 Dashboard-style Reporting ------------------------

@login_required
def reports(request):
    approved_count = ItemRequest.objects.filter(status='approved').count()
    pending_count = ItemRequest.objects.filter(status='pending').count()
    rejected_count = ItemRequest.objects.filter(status='rejected').count()
    issued_count = ItemRequest.objects.filter(status='issued').count()
    returned_count = ItemRequest.objects.filter(status='returned').count()

    available_count = InventoryItem.objects.filter(status='available').count()
    issued_items_count = InventoryItem.objects.filter(status='issued').count()
    maintenance_count = InventoryItem.objects.filter(status='maintenance').count()
    retired_count = InventoryItem.objects.filter(status='retired').count()

    context = {
        'approved_count': approved_count,
        'pending_count': pending_count,
        'rejected_count': rejected_count,
        'issued_count': issued_count,
        'returned_count': returned_count,
        'available_count': available_count,
        'issued_items_count': issued_items_count,
        'maintenance_count': maintenance_count,
        'retired_count': retired_count,
    }
    return render(request, 'reports/reports.html', context)


# ------------------------ 📄 General Item Request Report ------------------------

@login_required
def item_request_report(request):
    report_type = request.GET.get('report_type', 'monthly')
    year = request.GET.get('year')
    month = request.GET.get('month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    item = request.GET.get('item')
    status = request.GET.get('status')
    user = request.GET.get('user')

    queryset = ItemRequest.objects.all()
    selected_period = ""

    # Filter by time
    try:
        if report_type == 'yearly' and year:
            queryset = queryset.filter(request_date__year=year)
            selected_period = f"Year {year}"
        elif report_type == 'monthly' and month:
            year_val, month_val = map(int, month.split('-'))
            queryset = queryset.filter(request_date__year=year_val, request_date__month=month_val)
            selected_period = f"{calendar.month_name[month_val]} {year_val}"
        elif report_type == 'custom' and start_date and end_date:
            start = parse_date(start_date)
            end = parse_date(end_date)
            if start and end:
                queryset = queryset.filter(request_date__range=(start, end))
                selected_period = f"{start.strftime('%b %d, %Y')} – {end.strftime('%b %d, %Y')}"
    except Exception:
        selected_period = "Invalid range"

    # Filters
    if item:
        queryset = queryset.filter(item__name__icontains=item)
    if status:
        queryset = queryset.filter(status=status)
    if user:
        queryset = queryset.filter(user__username__icontains=user)

    status_choices = ['pending', 'approved', 'rejected', 'issued', 'returned']

    return render(request, 'reports/item_request_report.html', {
        'requests': queryset.order_by('-request_date'),
        'selected_period': selected_period,
        'status_choices': status_choices,
    })


# ------------------------ 📤 CSV Export for Requests ------------------------

@login_required
def export_request_report(request):
    format = request.GET.get('format', 'csv')
    month = request.GET.get('month')  # format: YYYY-MM
    item = request.GET.get('item')
    status = request.GET.get('status')
    user = request.GET.get('user')

    queryset = ItemRequest.objects.all()

    # Filter by month
    if month:
        try:
            year, month_num = map(int, month.split('-'))
            queryset = queryset.filter(request_date__year=year, request_date__month=month_num)
        except:
            return HttpResponseBadRequest("Invalid month format")

    if item:
        queryset = queryset.filter(item__name__icontains=item)
    if status:
        queryset = queryset.filter(status=status)
    if user:
        queryset = queryset.filter(user__username__icontains=user)

    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="item_request_report.csv"'

        writer = csv.writer(response)
        writer.writerow(['Item', 'User', 'Status', 'Request Date'])

        for req in queryset:
            writer.writerow([
                req.item.name,
                req.user.username,
                req.status,
                req.request_date.strftime("%Y-%m-%d %H:%M"),
            ])
        return response

    return HttpResponse("Invalid export format", status=400)


# ------------------------ 🧾 Inventory Report (Excel Export Optional) ------------------------

@login_required
def inventory_report(request):
    report_type = request.GET.get('report_type', 'monthly')
    year = request.GET.get('year')
    month = request.GET.get('month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    item = request.GET.get('item')

    queryset = InventoryItem.objects.all()
    selected_period = ""

    try:
        if report_type == 'yearly' and year:
            queryset = queryset.filter(date_added__year=int(year))
            selected_period = f"Year {year}"
        elif report_type == 'monthly' and month:
            year_val, month_val = map(int, month.split('-'))
            queryset = queryset.filter(date_added__year=year_val, date_added__month=month_val)
            selected_period = f"{calendar.month_name[month_val]} {year_val}"
        elif report_type == 'custom' and start_date and end_date:
            start = parse_date(start_date)
            end = parse_date(end_date)
            if start and end:
                queryset = queryset.filter(date_added__date__range=(start, end))
                selected_period = f"{start.strftime('%b %d, %Y')} – {end.strftime('%b %d, %Y')}"
        else:
            selected_period = "All Records"
    except Exception as e:
        selected_period = f"Invalid range: {e}"

    if item:
        queryset = queryset.filter(name__icontains=item)

    if request.GET.get('export') == 'excel':
        return export_inventory_to_excel(queryset, selected_period)

    return render(request, 'reports/inventory_report.html', {
        'items': queryset,
        'selected_period': selected_period,
    })


# ------------------------ 📥 Inventory Export to Excel ------------------------

def export_inventory_to_excel(queryset, selected_period):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    headers = ['#', 'Item Name', 'Serial Number', 'Category', 'Status', 'Location', 'Date Added']
    ws.append(headers)

    for idx, item in enumerate(queryset, start=1):
        ws.append([
            idx,
            item.name,
            item.serial_number,
            item.get_category_display() if hasattr(item, 'get_category_display') else item.category,
            item.status,
            item.location,
            item.date_added.strftime('%Y-%m-%d'),
        ])

    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventory_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ------------------------ 👤 User Dashboard Request Report ------------------------

@login_required
def user_item_request_report(request):
    report_type = request.GET.get('report_type', 'monthly')
    year = request.GET.get('year')
    month = request.GET.get('month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    item = request.GET.get('item')
    status = request.GET.get('status')

    queryset = ItemRequest.objects.filter(user=request.user)
    selected_period = ""

    # Filter by date
    try:
        if report_type == 'yearly' and year:
            queryset = queryset.filter(request_date__year=year)
            selected_period = f"Year {year}"
        elif report_type == 'monthly' and month:
            year_val, month_val = map(int, month.split('-'))
            queryset = queryset.filter(request_date__year=year_val, request_date__month=month_val)
            selected_period = f"{calendar.month_name[month_val]} {year_val}"
        elif report_type == 'custom' and start_date and end_date:
            start = parse_date(start_date)
            end = parse_date(end_date)
            if start and end:
                queryset = queryset.filter(request_date__range=(start, end))
                selected_period = f"{start.strftime('%b %d, %Y')} – {end.strftime('%b %d, %Y')}"
    except Exception:
        selected_period = "Invalid range"

    if item:
        queryset = queryset.filter(item__name__icontains=item)
    if status:
        queryset = queryset.filter(status=status)

    status_choices = ['pending', 'approved', 'rejected', 'issued', 'returned']

    return render(request, 'reports/item_request_report.html', {
        'requests': queryset.order_by('-request_date'),
        'selected_period': selected_period,
        'status_choices': status_choices,
    })
