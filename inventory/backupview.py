from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout, authenticate
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.http import JsonResponse, HttpResponseBadRequest
from django.contrib.auth.models import Group
from .utils import is_clerk, is_faculty, is_hod, forecast_usage_from_excel
from .models import ItemRequest, InventoryItem
from .forms import InventoryItemForm
from notifications.utils import notify_user
from django.contrib.auth.models import User
import datetime
import os
from django.core.mail import send_mail
from datetime import datetime
from django.core.paginator import Paginator


def home(request):
    if is_clerk(request.user):
        return redirect('clerk_dashboard')
    elif is_faculty(request.user):
        return redirect('faculty_dashboard')
    elif is_hod(request.user):
        return redirect('hod_dashboard')
    else:
        return redirect('login')
    
@login_required
def logout_page(request):
    logout(request)
    return redirect('login')

def login_page(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home') 
        else:
            messages.error(request, 'Invalid Credentials')
    return render(request, 'user/login.html')

@login_required
def view_all_requests(request):
    requests = ItemRequest.objects.all().order_by('-request_date')
    return render(request, 'requests/view_requests.html', {'requests': requests})

@login_required
def view_requests(request, user_id):
    requests = ItemRequest.objects.filter( user__id = user_id )
    return render(request, 'requests/view_requests.html', {'requests': requests})


@user_passes_test(is_hod or is_clerk)
def manage_requests(request):
    requests = ItemRequest.objects.all().order_by('-request_date')
    return render(request, 'inventory/manage_requests.html', {'requests': requests})

@user_passes_test(is_hod)
@require_POST
def process_request(request, request_id):
    item_request = get_object_or_404(ItemRequest, id=request_id)
    action = request.POST.get('action')

    if item_request.status != 'pending':
        messages.warning(request, "This request has already been processed.")
        return redirect('hod_dashboard')

    if action == 'approve':
        item = item_request.item
        if item.quantity >= item_request.quantity:
            item.quantity -= item_request.quantity
            item.save()

            item_request.status = 'approved'
            item_request.processed_by = request.user
            item_request.decision_date = datetime.now()
            item_request.save()
            notify_user(item_request.user, f"Your request for {item.name} has been approved by {request.user}.")
            clerk_group = Group.objects.get(name='Clerk')
            for clerk in clerk_group.user_set.all():
                notify_user(clerk, f" {item_request.user}'s request for {item.name} has been approved by {request.user}.")
            messages.success(request, "Request approved and inventory updated.")
        else:
            messages.error(request, "Not enough stock to approve this request.")
            return redirect('hod_dashboard')

    elif action == 'reject':
        item_request.status = 'rejected'
        item_request.processed_by = request.user
        item_request.save()
        notify_user(item_request.user, f"Your request for {item.item.name} has been rejected.")
        messages.info(request, "Request has been rejected.")

    else:
        messages.error(request, "Invalid action.")
    
    return redirect('hod_dashboard')


@login_required
def request_history(request):
    if is_faculty(request.user):
        requests = ItemRequest.objects.filter(user=request.user).order_by('-request_date')
    else: 
        requests = ItemRequest.objects.all().order_by('-request_date')
    return render(request, 'requests/request_history.html', {'requests': requests})


@login_required
@user_passes_test(is_clerk)
def clerk_dashboard(request):
    # total_items = InventoryItem.objects.count()
    # low_stock_items = InventoryItem.objects.filter(quantity__lt=5).count()  # Adjust threshold
    # pending_requests = ItemRequest.objects.filter(user=request.user, status='pending').count()
    # issued_items = ItemRequest.objects.filter( status='issued').count()

    recent_requests = ItemRequest.objects.filter().order_by('-request_date')[:5]
    recent_approved_requests = ItemRequest.objects.filter(status = 'approved').order_by('-request_date')[:5]

    context = {
        # 'total_items': total_items,
        # 'low_stock': low_stock_items,
        # 'pending_requests': pending_requests,
        # 'issued_items': issued_items,
        'recent_requests': recent_requests,
        'recent_approved_requests':recent_approved_requests
    }
    return render(request, 'dashboard/clerk_dashboard.html', context)

from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
@user_passes_test(is_hod)
def hod_dashboard(request):
    total_requests = ItemRequest.objects.count()
    pending_requests = ItemRequest.objects.filter(status='pending').count()
    approved_requests = ItemRequest.objects.filter(status='approved').count()
    rejected_requests = ItemRequest.objects.filter(status='rejected').count()
    issued_requests = ItemRequest.objects.filter(status='issued').count()
    returned_requests = ItemRequest.objects.filter(status='returned').count()

    recent_requests = ItemRequest.objects.all().order_by('-request_date')[:5]

    context = {
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
        'rejected_requests': rejected_requests,
        'issued_requests':issued_requests,
        'returned_requests':returned_requests,
        'recent_requests': recent_requests,
    }
    return render(request, 'dashboard/hod_dashboard.html', context)

@login_required
@user_passes_test(is_faculty)
def faculty_dashboard(request):
    items = InventoryItem.objects.all()
    my_requests = ItemRequest.objects.filter(user=request.user).order_by('-request_date')

    # Pagination for items
    items_paginator = Paginator(items, 100)  # Show 10 items per page
    items_page_number = request.GET.get('items_page')
    items_page = items_paginator.get_page(items_page_number)

    # Pagination for my_requests
    requests_paginator = Paginator(my_requests, 100)  # Show 10 requests per page
    requests_page_number = request.GET.get('requests_page')
    requests_page = requests_paginator.get_page(requests_page_number)

    return render(request, 'dashboard/faculty_dashboard.html', {
        'items': items_page,
        'my_requests': requests_page
    })

@login_required
@user_passes_test(is_clerk)
def issue_items(request):
    approved_requests = ItemRequest.objects.filter(status='approved').order_by('-request_date')
    return render(request, 'requests/issue_items.html', {'approved_requests': approved_requests})

@login_required
@user_passes_test(is_clerk)
@require_POST
def mark_as_issued(request, request_id):
    item_request = get_object_or_404(ItemRequest, id=request_id, status='approved')
    
    item = get_object_or_404(InventoryItem, id=item_request.item.id)
    request_quantity = item_request.quantity

    if item.quantity >= request_quantity:
        item.quantity -= request_quantity
        item.last_maintenance_date = datetime.now()
        item.save()
        item_request.status = 'issued'
        item_request.issued_date = datetime.now()
        item_request.save()
        messages.success(request, "Item marked as issued.")
    else:
        messages.error(request, "Not enough stock to issue the item.")

    return redirect('clerk_dashboard')  

@login_required
def show_all_items(request):
    items = InventoryItem.objects.all()

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    
    in_stock_items = items.filter(quantity__gte=5).count()   # Only items with good stock
    low_stock_items = items.filter(quantity__gt=0, quantity__lt=10).count()
    out_of_stock_items = items.filter(quantity=0).count()

    stock_info = {
        'in_stock': in_stock_items,
        'low_stock': low_stock_items,
        'out_stock': out_of_stock_items,
    }
    categories = InventoryItem.CATEGORY_CHOICES
    return render(request, 'inventory/show_all_items.html', {
        'items': page_obj,
        'stock_info': stock_info,
        'categories': categories,
    })

@login_required
@require_POST
@user_passes_test(is_clerk)
def delete_item(request, pk):
    try:
        item = InventoryItem.objects.get(pk=pk)
        item.delete()
        return JsonResponse({'success': True})
    except InventoryItem.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Item not found'})

@login_required
@user_passes_test(is_clerk)
def add_item(request):
    if request.method == 'POST':
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Item added successfully.")
            return redirect('add_item')
    else:
        form = InventoryItemForm()
    return render(request, 'inventory/add_edit_item.html', {'form': form})

@login_required
@user_passes_test(is_clerk)
def edit_item(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Item updated successfully.")
            return redirect('inventory_items')
    else:
        form = InventoryItemForm(instance=item)
    return render(request, 'inventory/add_edit_item.html', {'form': form, 'item': item})

@login_required
def request_item_page(request):
    items = InventoryItem.objects.filter(quantity__gt=0)
    return render(request, 'requests/request_item.html', {'items': items})

@login_required
def submit_item_request(request):
    if request.method == 'POST':
        item_id = request.POST.get('item_id')
        quantity = request.POST.get('quantity')
        reason = request.POST.get('reason')
        item = get_object_or_404(InventoryItem, id=item_id)

        ItemRequest.objects.create(
            user=request.user,
            item=item,
            quantity=quantity,
            reason=reason,
            status='pending',
            request_date = datetime.now()
        )
        hod_group = Group.objects.get(name='HOD')
        for hod in hod_group.user_set.all():
            notify_user(hod, f"{request.user.username} requested {item.name} of quantity {quantity}.")
        messages.success(request, 'Request submitted successfully.')
        return redirect('home')
    return HttpResponseBadRequest("Invalid request.")

# @login_required
# def category_list(request):
#     categories = InventoryItem.CATEGORY_CHOICES
#     if request.method == 'GET':
#         category = request.GET.get('category')
#         if category == 'all':
#             items = InventoryItem.objects.all()
#         else:
#             items = InventoryItem.objects.filter(category=category)
#     else:
#         items = InventoryItem.objects.all()
#     return render(request, 'inventory/show_all_items.html', {'categories': categories, 'items': items})

from django.shortcuts import render
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
import datetime
import os

def inventory_forecast_view(request):
    forecast_results = None
    forecast_year = [ datetime.datetime.now().year + 1, datetime.now().year + 2 ]  # Default to next year
    year = datetime.now().year + 1  
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        year = request.POST.get('forecast_year')
        if uploaded_file.name.endswith('.xlsx'):
            fs = FileSystemStorage()
            filename = fs.save(uploaded_file.name, uploaded_file)
            file_path = fs.path(filename)

            try:
                
                forecast_results = forecast_usage_from_excel(file_path, int(year))

            except Exception as e:
                messages.error(request, f"Error processing file: {e}")
            finally:
                os.remove(file_path)
        else:
            messages.error(request, "Please upload a .xlsx file only.")

    return render(request, 'forecast/forecast_page.html', {
        'forecast_data': forecast_results,
        'forecast_year': forecast_year,
        'selected_year': year
    })


@login_required
@user_passes_test(is_clerk or is_hod)
def predict_usage(request, year):
    forecast_data = {}
    try:
        year = int(year)
    except ValueError:
        return JsonResponse({'error': 'Invalid year format'}, status=400)
    forecast_year = year 

    excel_path = os.path.join('media', 'stationery_usage.xlsx') 
    try:
        forecast_data = forecast_usage_from_excel(excel_path, forecast_year)
    except Exception as e:
        forecast_data = {'error': str(e)}

    return render(request, 'predict_usage.html', {
        'forecast_data': forecast_data,
        'forecast_year': forecast_year
    })

@login_required
@user_passes_test(is_hod)
def manage_users(request):
    users = User.objects.all()
    return render(request, 'user/manage_users.html', {'users': users})


from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.contrib import messages
from django.template.loader import render_to_string
from .utils import is_hod

@login_required
@user_passes_test(is_hod)
def create_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        group_name = request.POST.get('group')
        email = request.POST.get('email')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists. Please choose a different one.")
            return redirect('create_user')

        try:
            validate_email(email)
        except ValidationError:
            messages.error(request, "Invalid email address provided.")
            return redirect('create_user')

        try:
            # Create user
            user = User.objects.create_user(username=username, password=password, email=email)

            # Assign group
            try:
                group = Group.objects.get(name=group_name)
                user.groups.add(group)
            except Group.DoesNotExist:
                user.delete()  # rollback
                messages.error(request, f"Group '{group_name}' does not exist.")
                return redirect('create_user')

            # Send welcome email
            subject = "Your Inventory System Account"
            message = render_to_string('emails/account_created_email.txt', {
                'username': username,
                'password': password,
            })

            send_mail(
                subject,
                message,
                'aniketwakte42@gmail.com',
                [email],
                fail_silently=False
            )

            messages.success(request, f"User {username} created successfully and notified by email.")
            return redirect('home')

        except Exception as e:
            messages.error(request, f"Error creating user: {e}")
            return redirect('create_user')

    return render(request, 'user/create_user.html')

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models.functions import TruncMonth
from django.db.models import Count
from .models import ItemRequest
import calendar
from django.utils.dateparse import parse_date
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from inventory.models import ItemRequest
from django.utils.dateparse import parse_date
from django.db.models import Q
import calendar

@login_required
def item_request_report(request):
    report_type = request.GET.get('report_type', 'monthly')
    year = request.GET.get('year')
    month = request.GET.get('month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    item = request.GET.get('item')
    status = request.GET.get('status')
    user = request.GET.get('user')

    queryset = ItemRequest.objects.none()
    selected_period = ""

    # Report type logic
    if report_type == 'yearly' and year:
        queryset = ItemRequest.objects.filter(request_date__year=year).order_by('request_date')
        selected_period = f"Year {year}"

    elif report_type == 'monthly' and month:
        try:
            year_val, month_val = map(int, month.split('-'))
            queryset = ItemRequest.objects.filter(request_date__year=year_val, request_date__month=month_val).order_by('request_date')
            selected_period = f"{calendar.month_name[month_val]} {year_val}"
        except Exception:
            selected_period = "Invalid month"

    elif report_type == 'custom' and start_date and end_date:
        try:
            start = parse_date(start_date)
            end = parse_date(end_date)
            if start and end:
                queryset = ItemRequest.objects.filter(request_date__range=(start, end)).order_by('request_date')
                selected_period = f"{start.strftime('%b %d, %Y')} - {end.strftime('%b %d, %Y')}"
        except Exception:
            selected_period = "Invalid custom range"
    # print(selected_period)
    # Apply filters
    if item:
        queryset = queryset.filter(item__name__icontains=item).order_by('request_date')
    if status:
        queryset = queryset.filter(status=status).order_by('request_date')
    if user:
        queryset = queryset.filter(user__username__icontains=user).order_by('request_date')

    status_choices = ['pending', 'approved', 'rejected', 'issued', 'returned']

    return render(request, 'reports/item_request_report.html', {
        'requests': queryset,
        'selected_period': selected_period,
        'status_choices': status_choices,
    })


import csv
from django.http import HttpResponse
from .models import ItemRequest


@login_required
def export_request_report(request):
    format = request.GET.get('format', 'csv')
    month = request.GET.get('month')
    item = request.GET.get('item')
    status = request.GET.get('status')
    user = request.GET.get('user')

    year, month_number = map(int, month.split('-'))
    queryset = ItemRequest.objects.filter(
        request_date__year=year,
        request_date__month=month_number
    )

    if item:
        queryset = queryset.filter(item__name__icontains=item)
    if status:
        queryset = queryset.filter(status=status)
    if user:
        queryset = queryset.filter(user__username__icontains=user)

    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="item_request_report.csv"'

        writer = csv.writer(response)
        writer.writerow(['Item', 'User', 'Quantity', 'Status', 'Request Date'])

        for req in queryset:
            writer.writerow([
                req.item.name,
                req.user.username,
                req.quantity,
                req.status,
                req.request_date.strftime("%Y-%m-%d %H:%M")
            ])
        return response

    # Future: handle PDF here

    return HttpResponse("Invalid format", status=400)


@login_required
def reports(request):
    approved_count = ItemRequest.objects.filter(status='approved').count()
    pending_count = ItemRequest.objects.filter(status='pending').count()
    rejected_count = ItemRequest.objects.filter(status='rejected').count()
    issued_count = ItemRequest.objects.filter(status='issued').count()
    returned_count = ItemRequest.objects.filter(status='returned').count()

    in_stock_count = InventoryItem.objects.filter(quantity__gt=10).count()
    low_stock_count = InventoryItem.objects.filter(quantity__gt=0, quantity__lte=10).count()
    out_of_stock_count = InventoryItem.objects.filter(quantity=0).count()

    context = {
        'approved_count': approved_count,
        'pending_count': pending_count,
        'rejected_count': rejected_count,
        'issued_count' : issued_count,
        'returned_count': returned_count,
        'in_stock_count': in_stock_count,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
    }

    return render(request, 'reports/reports.html', context)

from django.db.models import Q
from django.utils import timezone
from django.core.paginator import Paginator
import datetime

def item_maintenance(request):
    # Get search query and filter value from GET parameters
    search_query = request.GET.get('q', '')
    filter_value = request.GET.get('filter', '')

    # Start with the base query: items with 'issued' status
    issued_items = ItemRequest.objects.filter(status='issued').order_by('issued_date')
    issued_maintenance_items = issued_items.filter(item__needs_maintenance=True)
    # Filter by search query if provided
    if search_query:
        issued_maintenance_items = issued_maintenance_items.filter(Q(item__name__icontains=search_query))

    # Today's date
    today = timezone.now().date()

    # List to store filtered items
    filtered_items = []

    # Add custom logic to check if maintenance is overdue based on last_maintenance_date or issued_date
    for item in issued_maintenance_items:
        # Ensure last_maintenance_date is a date object
        last_maintenance_date = item.item.last_maintenance_date
        if last_maintenance_date and isinstance(last_maintenance_date, datetime.datetime):
            last_maintenance_date = last_maintenance_date.date()

        # Initialize is_overdue as False to avoid the "undefined variable" error
        is_overdue = False

        # If there's a last_maintenance_date
        if last_maintenance_date:
            # Calculate overdue based on last_maintenance_date
            if item.item.maintenance_interval_days is not None:
                days_since_last_maintenance = (today - last_maintenance_date).days
                is_overdue = days_since_last_maintenance > item.item.maintenance_interval_days
        else:
            # If there's no last_maintenance_date, calculate based on issued_date
            if item.item.maintenance_interval_days is not None:
                # Ensure issued_date is a date object (convert to date if it's a datetime object)
                # issued_date = item.issued_date.date() if isinstance(item.issued_date, datetime) else item.issued_date
                issued_date = item.issued_date.date() if isinstance(item.issued_date, datetime.datetime) else item.issued_date


                days_since_issued = (today - issued_date).days
                is_overdue = days_since_issued > item.item.maintenance_interval_days

        # Filter based on the filter_value ('due' or 'overdue')
        if filter_value == 'overdue' and is_overdue:
            filtered_items.append(item)
        elif filter_value == 'due' and not is_overdue:
            filtered_items.append(item)
        elif filter_value == '':  # Show all non-overdue items when no filter is set
            filtered_items.append(item)

    # Paginate the filtered items
    paginator = Paginator(filtered_items, 30)
    page_number = request.GET.get('items_page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'maintenance/maintenance_page.html', {'maintenance_items': page_obj})

import calendar
from django.utils.dateparse import parse_date

def inventory_report(request):
    report_type = request.GET.get('report_type', 'monthly')
    year = request.GET.get('year')
    month = request.GET.get('month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    item = request.GET.get('item')

    queryset = InventoryItem.objects.all()
    selected_period = ""

    try:
        if report_type == 'yearly' and year:
            year_int = int(year)
            queryset = queryset.filter(date_added__year=year_int)
            selected_period = f"Year {year_int}"
        elif report_type == 'monthly' and month:
            # Expecting month param in 'YYYY-MM' format
            year_val, month_val = map(int, month.split('-'))
            queryset = queryset.filter(date_added__year=year_val, date_added__month=month_val)
            selected_period = f"{calendar.month_name[month_val]} {year_val}"
        elif report_type == 'custom' and start_date and end_date:
            start = parse_date(start_date)
            end = parse_date(end_date)
            if start and end:
                queryset = queryset.filter(date_added__date__range=(start, end))
                selected_period = f"{start.strftime('%b %d, %Y')} - {end.strftime('%b %d, %Y')}"
            else:
                selected_period = "Invalid date range"
        else:
            # No valid filter - maybe show all or show message
            selected_period = "All time"
    except Exception as e:
        selected_period = f"Invalid filter parameters: {e}"

    if item:
        queryset = queryset.filter(name__icontains=item)

    if request.GET.get('export') == 'excel':
        return export_inventory_to_excel(queryset, selected_period)


    return render(request, 'reports/inventory_report.html', {
        'items': queryset,
        'selected_period': selected_period,
    })

import io
import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

def export_inventory_to_excel(queryset, selected_period):
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    # Header row
    headers = ['#', 'Item', 'Category', 'Quantity', 'Location', 'Date Added']
    ws.append(headers)

    # Fill rows
    for idx, item in enumerate(queryset, start=1):
        ws.append([
            idx,
            item.name,
            item.get_category_display() if hasattr(item, 'get_category_display') else item.category,
            item.quantity,
            item.location,
            item.date_added.strftime('%Y-%m-%d'),
        ])

    # Set column widths for readability
    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    # Save workbook to in-memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    filename = f"inventory_report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
