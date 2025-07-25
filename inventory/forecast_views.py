from datetime import datetime
import os
import csv
from io import BytesIO
from collections import Counter

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.utils.text import slugify

from .utils import forecast_usage_from_excel, is_clerk, is_hod, is_faculty
from inventory.models import ItemRequest, InventoryItem, ItemGroup
from statsmodels.tsa.holtwinters import ExponentialSmoothing


# # -------------------- 1. Manual Excel Upload Forecast View --------------------

# @login_required
# def inventory_forecast_view(request):
#     forecast_results = None
#     forecast_year = [datetime.now().year + 1, datetime.now().year + 2]
#     year = datetime.now().year + 1

#     if request.method == 'POST' and request.FILES.get('file'):
#         uploaded_file = request.FILES['file']
#         year = request.POST.get('forecast_year')

#         if uploaded_file.name.endswith('.xlsx'):
#             fs = FileSystemStorage()
#             filename = fs.save(uploaded_file.name, uploaded_file)
#             file_path = fs.path(filename)

#             try:
#                 forecast_results = forecast_usage_from_excel(file_path, int(year))
#             except Exception as e:
#                 messages.error(request, f"Error processing file: {e}")
#             finally:
#                 os.remove(file_path)
#         else:
#             messages.error(request, "Please upload a .xlsx file only.")

#     return render(request, 'forecast/forecast_page.html', {
#         'forecast_data': forecast_results,
#         'forecast_year': forecast_year,
#         'selected_year': year
#     })


# # -------------------- 2. API-based Prediction from Static Excel --------------------

# @login_required
# @user_passes_test(is_clerk or is_hod)
# def predict_usage(request, year):
#     forecast_data = {}
#     try:
#         year = int(year)
#     except ValueError:
#         return JsonResponse({'error': 'Invalid year format'}, status=400)

#     excel_path = os.path.join('media', 'stationery_usage.xlsx')
#     try:
#         forecast_data = forecast_usage_from_excel(excel_path, year)
#     except Exception as e:
#         forecast_data = {'error': str(e)}

#     return render(request, 'predict_usage.html', {
#         'forecast_data': forecast_data,
#         'forecast_year': year
#     })


# # -------------------- 3. System-based Forecast with Visualization --------------------

# @login_required
# def forecast_inventory_usage(request, forecast_years=1):
#     # Get all issued requests for consumables assigned to a group
#     issued_requests = ItemRequest.objects.filter(
#         status='issued',
#         item__item_type='consumable',
#         item__item_group__isnull=False
#     ).select_related('item__item_group').order_by('request_date')

#     if not issued_requests.exists():
#         return render(request, 'predict_usage2.html', {
#             'error': 'No issued item requests found.',
#             'forecast_data': {},
#             'past_usage': {},
#             'current_stock': {},
#             'need_to_order': {},
#             'forecast_year': datetime.now().year + 1,
#             'years_to_show': [],
#             'forecast_years_used': [],
#             'group_mode': True,
#         })

#     current_year = datetime.now().year
#     years_to_show = list(range(current_year - 4, current_year))

#     # Get all item groups in your system
#     item_groups = ItemGroup.objects.all()

#     # Initialize usage_data dict by group and year
#     usage_data = {group.name: {year: 0 for year in years_to_show} for group in item_groups}

#     # Count available items by group to get current stock per group
#     current_stock = {}
#     for group in item_groups:
#         qty = InventoryItem.objects.filter(
#             item_type='consumable',
#             status='available',
#             item_group=group
#         ).count()
#         current_stock[group.name] = qty

#     # Aggregate usage count of issued items per group per year
#     for req in issued_requests:
#         group = req.item.item_group
#         if group and req.request_date.year in usage_data[group.name]:
#             usage_data[group.name][req.request_date.year] += 1  # one unit per issued item

#     forecast_result = {}
#     need_to_order = {}

#     for group_name, yearly_data in usage_data.items():
#         sorted_years = sorted(yearly_data.keys())
#         values = [yearly_data[year] for year in sorted_years]
#         idx = pd.PeriodIndex(sorted_years, freq='Y')
#         series = pd.Series(values, index=idx)

#         if series.sum() == 0 or series.count() < 3:
#             # Insufficient data for forecasting
#             forecast_result[group_name] = "Not enough data"
#             need_to_order[group_name] = "-"
#             continue

#         try:
#             model = ExponentialSmoothing(series, trend='add', seasonal=None)
#             fit = model.fit()
#             forecast = fit.forecast(1)
#             forecast_year = str(current_year + 1)
#             forecast_value = int(round(forecast.iloc[0]))

#             forecast_result[group_name] = {forecast_year: forecast_value}

#             available_qty = current_stock.get(group_name, 0)
#             need_qty = forecast_value - available_qty
#             need_to_order[group_name] = need_qty if need_qty > 0 else 0

#         except Exception as e:
#             forecast_result[group_name] = f"Forecasting error: {str(e)}"
#             need_to_order[group_name] = "-"

#     return render(request, 'predict_usage2.html', {
#         'forecast_data': forecast_result,
#         'past_usage': usage_data,
#         'current_stock': current_stock,
#         'need_to_order': need_to_order,
#         'forecast_year': current_year + 1,
#         'years_to_show': years_to_show,
#         'forecast_years_used': [str(current_year + 1)],
#         'group_mode': True,
#     })


# # -------------------- 4. Generate Official Forecast Excel --------------------
# @csrf_exempt
# @login_required
# def generate_forecast_excel_report(request):
#     from collections import Counter
#     import openpyxl
#     from openpyxl.styles import Alignment, Border, Side
#     from openpyxl.utils import get_column_letter
#     from io import BytesIO
#     from datetime import datetime
#     from django.http import HttpResponse

#     if request.method == 'POST':
#         item_groups = ItemGroup.objects.all()
#         available_items = InventoryItem.objects.filter(item_type='consumable', status='available', item_group__isnull=False)
#         current_stock = dict(Counter(item.item_group.name for item in available_items if item.item_group))

#         item_data = []
#         for group in item_groups:
#             group_name = group.name
#             field_name = f"order_qty_{group_name.lower().replace(' ', '-').replace('/', '-')}"
#             order_qty = request.POST.get(field_name)
#             try:
#                 order_qty = int(order_qty)
#             except:
#                 order_qty = 0

#             # Most recent issued request for this group
#             latest_request = ItemRequest.objects.filter(item__item_group=group, status='issued').order_by('-request_date').first()
#             prev_date = latest_request.request_date.strftime('%d-%m-%Y') if latest_request else "N/A"

#             balance_quantity = current_stock.get(group_name, 0)
#             forecasted_qty = order_qty + balance_quantity
#             prev_qty = forecasted_qty - order_qty

#             item_data.append({
#                 "name": group_name,
#                 "sanctioned": forecasted_qty,
#                 "prev_qty": prev_qty,
#                 "prev_date": prev_date,
#                 "balance": balance_quantity,
#                 "required": order_qty,
#             })

#         # Generate Excel
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Stationery Forecast Report"
#         headers = [
#             "Sr. No.", "Stationery Item Group", "Sanctioned Qty",
#             "Previous received qty. & Date", "Balance Qty.", "Current required Qty."
#         ]
#         ws.append(headers)
#         thin = Side(border_style="thin", color="000000")

#         for i, item in enumerate(item_data, start=1):
#             row = [
#                 i,
#                 item["name"],
#                 item["sanctioned"],
#                 f"{item['prev_qty']} ({item['prev_date']})",
#                 item["balance"],
#                 item["required"]
#             ]
#             ws.append(row)
#             for col in range(1, 7):
#                 cell = ws.cell(row=ws.max_row, column=col)
#                 cell.alignment = Alignment(horizontal="center", vertical="center")
#                 cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
#         for col in ws.columns:
#             max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
#             ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)
#         filename = f"stationery_forecast_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
#         response = HttpResponse(
#             output,
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         return response

#     return HttpResponse("Invalid request method", status=400)


@login_required
def forecast_dashboard(request):
    """
    Renders group-level forecast, current stock and order input.
    GET   display
    POST  save Excel
    """
    from .forecast_views import build_forecast  # helper (section 2.3)

    # ❶ build all dicts (past_usage, forecast, stock, need_to_order)
    ctx = build_forecast(years_back=5, horizon=1)

    # ❷ Get years for template
    years = []
    if ctx.get('past_usage'):
        years = list(sorted(next(iter(ctx['past_usage'].values())).keys()))
    ctx["years"] = years

    # ❸ POST → Excel export
    if request.method == "POST":
        from .forecast_views import export_forecast_excel  # helper (section 2.4)
        return export_forecast_excel(request, ctx)

    return render(request, "forecast/dashboard.html", ctx)


@login_required
def forecast_api(request, group_id):
    """
    Returns JSON with past usage & forecast for one ItemGroup.
    """
    group = get_object_or_404(ItemGroup, id=group_id)
    data = build_forecast(single_group=group, years_back=5, horizon=1)
    return JsonResponse(data, safe=False)

def build_forecast(years_back:int=5, horizon:int=1, single_group=None):
    """
    Returns:
      forecast_data {grp:{yr:qty}}, past_usage {grp:{yr:qty}},
      current_stock {}, need_to_order {}
    """
    current_year = datetime.now().year
    years = list(range(current_year-years_back, current_year))
    groups = ItemGroup.objects.all() if single_group is None else [single_group]

    # Initialise dicts
    past_usage = {g.name:{y:0 for y in years} for g in groups}
    stock = {g.name:0 for g in groups}
    
    # Stock count
    qs_stock = InventoryItem.objects.filter(item_type="consumable", item_group__in=groups)
    for itm in qs_stock:
        stock[itm.item_group.name] += 1

    # Usage count
    qs_issued = (ItemRequest.objects
                 .filter(status="issued",
                         item__item_group__in=groups,
                         item__item_type="consumable"))
    for req in qs_issued:
        yr = req.request_date.year
        if yr in years:
            past_usage[req.item.item_group.name][yr] += 1

    # Forecast loop
    forecast, need = {}, {}
    for g in groups:
        series = pd.Series(
            [past_usage[g.name][y] for y in years],
            index=pd.PeriodIndex(years, freq="Y")
        )

        if series.sum() == 0 or series.count() < 3:
            forecast[g.name] = "Not enough data"
            need[g.name] = "-"
            continue

        # Auto-model selection
        best_model, best_aic = None, float("inf")
        candidates = {
            "HW": lambda s: ExponentialSmoothing(s, trend="add").fit(),
            "SES": lambda s: ExponentialSmoothing(s, trend=None).fit(),
        }
        for name, builder in candidates.items():
            try:
                m = builder(series)
                if m.aic < best_aic:
                    best_model, best_aic = m, m.aic
            except Exception:  # skip bad model
                continue

        if not best_model:
            forecast[g.name] = "Forecast error"
            need[g.name]     = "-"
            continue

        pred_qty = int(round(best_model.forecast(horizon).iloc[-1]))
        next_year = str(current_year+1)
        forecast[g.name] = {next_year: pred_qty}

        diff = pred_qty - stock[g.name]
        need[g.name] = diff if diff > 0 else 0
        

    return dict(
        forecast_data = forecast,
        past_usage    = past_usage,
        current_stock = stock,
        need_to_order = need,
        forecast_years_used = [str(current_year+1)],
    )

def export_forecast_excel(request, ctx):
    """
    Reads order_qty_* from POST, merges with ctx, returns XLSX.
    """
    wb, ws = openpyxl.Workbook(), None
    ws = wb.active; ws.title = "Consumable Forecast"
    headers = ["#", "Item Group", "Current Stock", "Forecast Qty",
               "Suggested Order", "Entered Order"]
    ws.append(headers)

    thin = Side(border_style="thin", color="000")
    row_no = 1
    for grp, fcast in ctx['forecast_data'].items():
        row_no += 1
        forecast_val = list(fcast.values())[0] if isinstance(fcast, dict) else "-"
        suggest      = ctx['need_to_order'].get(grp, "-")
        input_name   = f"order_qty_{slugify(grp)}"
        entered      = request.POST.get(input_name, "0")
        ws.append([row_no-1, grp, ctx['current_stock'][grp],
                   forecast_val, suggest, entered])

        for col in range(1, 7):
            c = ws.cell(row=row_no, column=col)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # column width
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"forecast_{datetime.now():%Y%m%d_%H%M}.xlsx"
    resp = HttpResponse(buf,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
